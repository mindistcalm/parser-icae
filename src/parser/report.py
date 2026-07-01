from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from parser.config import AppConfig, load_config
from parser.models import Mention


MONTH_NAMES_RU = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}


def format_date(dt: datetime | None) -> str:
    if dt is None:
        return "не указана"
    return dt.strftime("%d.%m.%Y")


class ReportGenerator:
    def __init__(self, config: AppConfig | None = None) -> None:
        self.config = config or load_config()

    def generate(
        self,
        mentions: list[Mention],
        year: int,
        month: int,
    ) -> tuple[Path, Path]:
        output_dir = Path(self.config.reports.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        month_name = MONTH_NAMES_RU[month]
        base_name = f"icae_mentions_{year:04d}_{month:02d}"

        xlsx_path = output_dir / f"{base_name}.xlsx"
        html_path = output_dir / f"{base_name}.html"

        self._write_xlsx(mentions, xlsx_path, year, month, month_name)
        self._write_html(mentions, html_path, year, month, month_name)

        return xlsx_path, html_path

    def _write_xlsx(
        self,
        mentions: list[Mention],
        path: Path,
        year: int,
        month: int,
        month_name: str,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name} {year}"

        headers = [
            "№",
            "Ресурс",
            "Заголовок",
            "Ссылка",
            "Дата публикации",
            "Источник поиска",
        ]
        ws.append(headers)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, m in enumerate(mentions, start=1):
            ws.append(
                [
                    idx,
                    m.source_name,
                    m.title,
                    m.url,
                    format_date(m.published_at),
                    m.source_type.value,
                ]
            )

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 14

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(path)

    def _write_html(
        self,
        mentions: list[Mention],
        path: Path,
        year: int,
        month: int,
        month_name: str,
    ) -> None:
        org = self.config.organization
        rows = []
        for idx, m in enumerate(mentions, start=1):
            rows.append(
                f"""
                <tr>
                  <td>{idx}</td>
                  <td>{_esc(m.source_name)}</td>
                  <td>{_esc(m.title)}</td>
                  <td><a href="{_esc(m.url)}" target="_blank">{_esc(m.url)}</a></td>
                  <td>{format_date(m.published_at)}</td>
                </tr>
                """
            )

        html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>Упоминания {org.short_name} — {month_name} {year}</title>
  <style>
    body {{ font-family: system-ui, sans-serif; margin: 2rem; color: #1a1a1a; }}
    h1 {{ font-size: 1.4rem; }}
    .meta {{ color: #555; margin-bottom: 1.5rem; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 0.5rem 0.75rem; vertical-align: top; }}
    th {{ background: #f3f4f6; text-align: left; }}
    tr:nth-child(even) {{ background: #fafafa; }}
    a {{ color: #1d4ed8; word-break: break-all; }}
  </style>
</head>
<body>
  <h1>Упоминания {org.full_name} ({org.city})</h1>
  <p class="meta">Период: {month_name} {year}. Найдено записей: {len(mentions)}.</p>
  <table>
    <thead>
      <tr>
        <th>№</th>
        <th>Ресурс</th>
        <th>Заголовок</th>
        <th>Ссылка</th>
        <th>Дата публикации</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows) if rows else '<tr><td colspan="5">Упоминаний не найдено</td></tr>'}
    </tbody>
  </table>
</body>
</html>
"""
        path.write_text(html, encoding="utf-8")


def _esc(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
